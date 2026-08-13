"""In-memory XLSX parsing and validation for bulk user creation."""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from yuxi.services.rbac_service import (
    BUILTIN_ROLE_DEFINITIONS,
    SCOPE_RANK,
    get_user_permission_map,
)
from yuxi.services.user_identity_service import is_valid_phone_number, validate_username
from yuxi.storage.postgres.models_business import (
    Department,
    RBACPermission,
    RBACRole,
    RBACRolePermission,
    User,
)

MAX_IMPORT_FILE_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 500
IMPORT_COLUMNS = ("uid", "username", "phone_number", "password", "department", "roles")
REQUIRED_COLUMNS = {"uid", "username", "password", "department"}
WEAK_PASSWORDS = {"password", "password123", "12345678", "qwerty123", "admin123"}


def create_user_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "用户导入"
    sheet.append(IMPORT_COLUMNS)
    sheet.append(("zhangsan", "张三", "13800138000", "ChangeMe123", "信息部", "普通用户"))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="168CA3")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F2"
    widths = (20, 18, 18, 20, 18, 35)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    notes = workbook.create_sheet("填写说明")
    notes.append(("字段", "要求"))
    notes.append(("uid", "必填，3-20 位字母、数字或下划线，作为登录 ID"))
    notes.append(("username", "必填，2-20 位中文、字母、数字或下划线"))
    notes.append(("phone_number", "选填，中国大陆手机号；填写时必须唯一"))
    notes.append(("password", "必填，至少 8 位，且同时包含字母和数字"))
    notes.append(("department", "必填，填写已有部门名称或 ID"))
    notes.append(("roles", "选填，角色名称或代码；多个角色用逗号分隔，默认普通用户"))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _split_roles(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[,，;；]", value) if item.strip()]


def _valid_password(value: str) -> bool:
    return (
        len(value) >= 8
        and value.lower() not in WEAK_PASSWORDS
        and re.search(r"[A-Za-z]", value) is not None
        and re.search(r"\d", value) is not None
    )


def _read_rows(content: bytes, filename: str | None) -> list[dict[str, str]]:
    if not filename or not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持 .xlsx 文件")
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件为空")
    if len(content) > MAX_IMPORT_FILE_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="导入文件不能超过 5 MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        raw_headers = next(values, None)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="无法读取 XLSX 文件") from exc
    if not raw_headers:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件缺少表头")
    headers = [_cell_text(value).lower() for value in raw_headers]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"缺少必填列：{', '.join(sorted(missing))}",
        )
    rows: list[dict[str, str]] = []
    for values_row in values:
        row = {header: _cell_text(value) for header, value in zip(headers, values_row, strict=False) if header}
        if any(row.values()):
            rows.append(row)
        if len(rows) > MAX_IMPORT_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"单次最多导入 {MAX_IMPORT_ROWS} 名用户",
            )
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="导入文件没有用户数据")
    return rows


async def validate_user_import(
    db: AsyncSession,
    content: bytes,
    filename: str | None,
    *,
    current_user: User,
    permission_scope: str,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    raw_rows = _read_rows(content, filename)
    department_result = await db.execute(select(Department))
    departments = list(department_result.scalars().all())
    departments_by_name = {department.name.strip().lower(): department for department in departments}
    departments_by_id = {str(department.id): department for department in departments}
    role_result = await db.execute(select(RBACRole))
    roles = list(role_result.scalars().all())
    roles_by_key: dict[str, list[RBACRole]] = {}
    for role in roles:
        roles_by_key.setdefault(role.code.strip().lower(), []).append(role)
        roles_by_key.setdefault(role.name.strip().lower(), []).append(role)
    member_role = next(
        (role for role in roles if role.code == BUILTIN_ROLE_DEFINITIONS["user"]["code"]),
        None,
    )
    if member_role is None:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="系统基础角色尚未初始化")
    actor_permissions = await get_user_permission_map(db, current_user)
    role_permission_result = await db.execute(
        select(RBACRolePermission.role_id, RBACPermission.code, RBACRolePermission.scope).join(
            RBACPermission,
            RBACPermission.id == RBACRolePermission.permission_id,
        )
    )
    grants_by_role: dict[int, dict[str, str]] = {}
    for role_id, code, scope in role_permission_result.all():
        grants_by_role.setdefault(role_id, {})[code] = scope

    uids = [_cell_text(row.get("uid")).lower() for row in raw_rows if _cell_text(row.get("uid"))]
    phones = [_cell_text(row.get("phone_number")) for row in raw_rows if _cell_text(row.get("phone_number"))]
    usernames = [_cell_text(row.get("username")) for row in raw_rows if _cell_text(row.get("username"))]
    existing_result = await db.execute(
        select(User.uid, User.username, User.phone_number).where(
            or_(
                func.lower(User.uid).in_(uids),
                User.username.in_(usernames),
                User.phone_number.in_(phones),
            )
        )
    )
    existing_uids: set[str] = set()
    existing_usernames: set[str] = set()
    existing_phones: set[str] = set()
    for uid, username, phone in existing_result.all():
        existing_uids.add(uid.lower())
        existing_usernames.add(username)
        if phone:
            existing_phones.add(phone)

    uid_counts = {value: uids.count(value) for value in set(uids)}
    username_counts = {value: usernames.count(value) for value in set(usernames)}
    phone_counts = {value: phones.count(value) for value in set(phones)}
    output_rows: list[dict[str, Any]] = []
    for row_number, raw in enumerate(raw_rows, 2):
        uid = _cell_text(raw.get("uid")).lower()
        username = _cell_text(raw.get("username"))
        phone = _cell_text(raw.get("phone_number")) or None
        password = _cell_text(raw.get("password"))
        department_value = _cell_text(raw.get("department"))
        errors: list[str] = []
        if not re.fullmatch(r"[a-zA-Z0-9_]{3,20}", uid):
            errors.append("uid 必须为 3-20 位字母、数字或下划线")
        elif uid_counts.get(uid, 0) > 1:
            errors.append("uid 在文件内重复")
        elif uid in existing_uids:
            errors.append("uid 已存在")
        valid_username, username_error = validate_username(username)
        if not valid_username:
            errors.append(username_error)
        elif username_counts.get(username, 0) > 1:
            errors.append("用户名在文件内重复")
        elif username in existing_usernames:
            errors.append("用户名已存在")
        if phone:
            if not is_valid_phone_number(phone):
                errors.append("手机号格式不正确")
            elif phone_counts.get(phone, 0) > 1:
                errors.append("手机号在文件内重复")
            elif phone in existing_phones:
                errors.append("手机号已存在")
        if not _valid_password(password):
            errors.append("密码至少 8 位并同时包含字母和数字，且不能使用常见弱密码")

        target_department = departments_by_id.get(department_value) or departments_by_name.get(
            department_value.lower()
        )
        if target_department is None:
            errors.append("部门不存在")
        elif permission_scope != "global" and target_department.id != current_user.department_id:
            errors.append("无权在其他部门创建用户")

        role_values = _split_roles(_cell_text(raw.get("roles")))
        selected_roles: list[RBACRole] = []
        for role_value in role_values:
            candidates = roles_by_key.get(role_value.lower(), [])
            role = next(
                (
                    candidate
                    for candidate in candidates
                    if target_department is not None and candidate.department_id == target_department.id
                ),
                next((candidate for candidate in candidates if candidate.department_id is None), None),
            )
            if role is None:
                errors.append(f"角色不存在：{role_value}")
            elif role not in selected_roles:
                selected_roles.append(role)
        if not selected_roles:
            selected_roles = [member_role]
        if target_department is not None:
            for role in selected_roles:
                if role.department_id is not None and role.department_id != target_department.id:
                    errors.append(f"角色不属于目标部门：{role.name}")
        base_roles = [role for role in selected_roles if role.is_system]
        if not base_roles:
            selected_roles.insert(0, member_role)
            base_roles = [member_role]
        if len(base_roles) > 1:
            errors.append("只能选择一个系统基础角色")
        if any(role.code == BUILTIN_ROLE_DEFINITIONS["superadmin"]["code"] for role in selected_roles):
            errors.append("不能批量创建超级管理员")
        if current_user.role != "superadmin" and any(
            role.code != BUILTIN_ROLE_DEFINITIONS["user"]["code"] for role in base_roles
        ):
            errors.append("部门管理员只能创建普通用户")
        for role in selected_roles:
            if role.is_system:
                continue
            if role.department_id is None and actor_permissions.get("role.assign") != "global":
                errors.append(f"无权分配公司级角色：{role.name}")
                continue
            if any(
                SCOPE_RANK.get(actor_permissions.get(code), 0) < SCOPE_RANK.get(scope, 0)
                for code, scope in grants_by_role.get(role.id, {}).items()
            ):
                errors.append(f"角色权限超出操作者授权范围：{role.name}")

        output_rows.append(
            {
                "row": row_number,
                "status": "invalid" if errors else "valid",
                "errors": errors,
                "uid": uid,
                "username": username,
                "phone_number": phone,
                "department_id": target_department.id if target_department else None,
                "department_name": target_department.name if target_department else department_value,
                "role_ids": [role.id for role in selected_roles],
                "role_names": [role.name for role in selected_roles],
                "_password": password,
            }
        )
    invalid = sum(row["status"] == "invalid" for row in output_rows)
    return {
        "total": len(output_rows),
        "valid": len(output_rows) - invalid,
        "invalid": invalid,
    }, output_rows


def public_import_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{key: value for key, value in row.items() if key != "_password"} for row in rows]
