from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from yuxi.services.user_import_service import (
    IMPORT_COLUMNS,
    _read_rows,
    create_user_import_template,
    public_import_rows,
    validate_user_import,
)
from yuxi.storage.postgres.models_business import (
    Department,
    RBACPermission,
    RBACRole,
    RBACRolePermission,
    RBACUserRole,
    User,
)


def _xlsx(*rows: tuple[str, ...]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(IMPORT_COLUMNS)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_template_contains_stable_columns_and_instructions():
    workbook = load_workbook(BytesIO(create_user_import_template()), read_only=True)
    assert workbook.sheetnames == ["用户导入", "填写说明"]
    assert tuple(next(workbook["用户导入"].iter_rows(values_only=True))) == IMPORT_COLUMNS


def test_read_rows_rejects_non_xlsx_and_missing_required_column():
    with pytest.raises(HTTPException, match="仅支持"):
        _read_rows(b"data", "users.csv")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("uid", "username"))
    sheet.append(("u001", "用户"))
    output = BytesIO()
    workbook.save(output)
    with pytest.raises(HTTPException, match="缺少必填列"):
        _read_rows(output.getvalue(), "users.xlsx")


def test_read_rows_rejects_spoofed_xlsx():
    with pytest.raises(HTTPException, match="无法读取"):
        _read_rows(b"not a zip workbook", "users.xlsx")


def test_public_rows_never_expose_password():
    rows = [{"row": 2, "uid": "u001", "_password": "DoNotReturn123", "status": "valid"}]
    assert public_import_rows(rows) == [{"row": 2, "uid": "u001", "status": "valid"}]


def test_read_rows_applies_row_limit(monkeypatch):
    monkeypatch.setattr("yuxi.services.user_import_service.MAX_IMPORT_ROWS", 1)
    content = _xlsx(
        ("u001", "用户甲", "", "Pass1234", "信息部", ""),
        ("u002", "用户乙", "", "Pass1234", "信息部", ""),
    )
    with pytest.raises(HTTPException, match="单次最多导入"):
        _read_rows(content, "users.xlsx")


@pytest.mark.asyncio
async def test_validation_reports_duplicate_uid_unknown_department_and_role():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as connection:
        await connection.run_sync(lambda sync_connection: Department.__table__.create(sync_connection))
        await connection.run_sync(lambda sync_connection: User.__table__.create(sync_connection))
        await connection.run_sync(lambda sync_connection: RBACPermission.__table__.create(sync_connection))
        await connection.run_sync(lambda sync_connection: RBACRole.__table__.create(sync_connection))
        await connection.run_sync(lambda sync_connection: RBACRolePermission.__table__.create(sync_connection))
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    async with session_factory() as db:
        department = Department(name="信息部")
        member_role = RBACRole(code="system.member", name="普通用户", is_system=True)
        db.add_all([department, member_role])
        await db.commit()
        actor = User(
            id=999,
            uid="superadmin",
            username="超级管理员",
            password_hash="hash",
            role="superadmin",
            department_id=department.id,
        )
        content = _xlsx(
            ("same_uid", "用户甲", "13800138000", "Secure123", "信息部", "普通用户"),
            ("same_uid", "用户乙", "13900139000", "Secure123", "不存在部门", "不存在角色"),
        )
        summary, rows = await validate_user_import(
            db,
            content,
            "users.xlsx",
            current_user=actor,
            permission_scope="global",
        )
        public_rows = public_import_rows(rows)
        assert summary == {"total": 2, "valid": 0, "invalid": 2}
        assert "uid 在文件内重复" in public_rows[0]["errors"]
        assert "部门不存在" in public_rows[1]["errors"]
        assert "角色不存在：不存在角色" in public_rows[1]["errors"]
        assert all("_password" not in row and "password" not in row for row in public_rows)
    await engine.dispose()


@pytest.mark.asyncio
async def test_validation_rejects_cross_department_and_elevated_roles():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as connection:
        for table in (
            Department.__table__,
            User.__table__,
            RBACPermission.__table__,
            RBACRole.__table__,
            RBACRolePermission.__table__,
            RBACUserRole.__table__,
        ):
            await connection.run_sync(lambda sync_connection, table=table: table.create(sync_connection))
    session_factory = async_sessionmaker(engine, expire_on_commit=False)
    async with session_factory() as db:
        own_department = Department(name="信息部")
        other_department = Department(name="财务部")
        member_role = RBACRole(code="system.member", name="普通用户", is_system=True)
        admin_role = RBACRole(code="system.department_admin", name="部门管理员", is_system=True)
        db.add_all([own_department, other_department, member_role, admin_role])
        await db.commit()
        actor = User(
            id=999,
            uid="department_admin",
            username="部门管理员",
            password_hash="hash",
            role="admin",
            department_id=own_department.id,
        )
        content = _xlsx(
            ("member001", "成员甲", "", "Secure123", "财务部", "普通用户"),
            ("member002", "成员乙", "", "Secure123", "信息部", "部门管理员"),
        )
        summary, rows = await validate_user_import(
            db,
            content,
            "users.xlsx",
            current_user=actor,
            permission_scope="department",
        )
        assert summary == {"total": 2, "valid": 0, "invalid": 2}
        assert "无权在其他部门创建用户" in rows[0]["errors"]
        assert "部门管理员只能创建普通用户" in rows[1]["errors"]
    await engine.dispose()
